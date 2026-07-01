from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.utils import timezone
from .models import Role, PercorsoCrescita, AuditLog, UserDocument, RichiestaCreazioneProfilo, RichiestaEliminazioneProfilo, NotificaTrasferimento
from participants.models import Iscrizione
from .serializers import (
    CustomTokenObtainPairSerializer, UserListSerializer,
    UserDetailSerializer, RoleSerializer, PercorsoCrescitaSerializer,
    AuditLogSerializer, RichiestaCreazionoProfiloSerializer,
    RichiestaEliminazioneProfiloSerializer,
)
from .permissions import IsAdmin, IsAdminOrHO, IsAdminOrHOOrArea, CanManageUsers


TIPO_LABEL_AUDIT = {
    'promozione': 'Promozione',
    'trasferimento': 'Trasferimento',
    'cambio_ruolo': 'Cambio Ruolo',
    'nota': 'Nota',
}


def log_audit(actor, azione, target_tipo, target_obj, dettaglio=''):
    try:
        AuditLog.objects.create(
            actor=actor,
            azione=azione,
            target_tipo=target_tipo,
            target_id=target_obj.pk if target_obj else None,
            target_repr=str(target_obj) if target_obj else '',
            dettaglio=dettaglio,
        )
    except Exception:
        pass

User = get_user_model()


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data['refresh']
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response({'detail': 'Logout effettuato.'})
        except Exception:
            return Response({'detail': 'Token non valido.'}, status=status.HTTP_400_BAD_REQUEST)


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserDetailSerializer(request.user)
        return Response(serializer.data)

    def patch(self, request):
        serializer = UserDetailSerializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.validated_data.pop('livello_accesso', None)
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserListCreateView(generics.ListCreateAPIView):
    permission_classes = [CanManageUsers]
    pagination_class = None

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserDetailSerializer
        return UserListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = User.objects.select_related('store', 'ruolo').filter(is_active=True)
        if user.livello_accesso == 'store':
            qs = qs.filter(store=user.store)
        store_id = self.request.query_params.get('store')
        if store_id:
            qs = qs.filter(store_id=store_id)
        return qs

    def perform_create(self, serializer):
        from rest_framework.exceptions import PermissionDenied, ValidationError
        livello_chi_crea = self.request.user.livello_accesso
        if livello_chi_crea not in ('admin', 'ho'):
            raise PermissionDenied('Solo Admin e HO possono creare utenti.')
        livello_nuovo = serializer.validated_data.get('livello_accesso', 'base')
        if livello_chi_crea in ('admin', 'ho') and livello_nuovo not in ('area', 'base'):
            raise ValidationError({'livello_accesso': 'Admin e HO possono creare solo profili Area Manager o Base.'})
        if not serializer.validated_data.get('store'):
            raise ValidationError({'store': 'Il workplace è obbligatorio.'})
        instance = serializer.save()
        log_audit(self.request.user, 'crea', 'utente', instance,
                  f'Nuovo utente creato: {instance.email}')


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [CanManageUsers]
    serializer_class = UserDetailSerializer

    def get_queryset(self):
        user = self.request.user
        if user.livello_accesso in ('admin', 'ho'):
            return User.objects.select_related('store', 'ruolo').all()
        if user.livello_accesso == 'store':
            return User.objects.filter(store=user.store)
        return User.objects.none()

    def perform_update(self, serializer):
        from rest_framework.exceptions import ValidationError
        livello_chi_modifica = self.request.user.livello_accesso
        nuovo_livello = serializer.validated_data.get('livello_accesso')
        if nuovo_livello and livello_chi_modifica in ('admin', 'ho') and nuovo_livello not in ('area', 'base'):
            raise ValidationError({'livello_accesso': 'Admin e HO possono assegnare solo livelli Area Manager o Base.'})
        if 'store' in serializer.validated_data and not serializer.validated_data.get('store'):
            raise ValidationError({'store': 'Il workplace è obbligatorio.'})

        FIELD_LABELS = {
            'cognome': 'Cognome', 'nome': 'Nome', 'email': 'Email',
            'sesso': 'Sesso', 'livello_accesso': 'Livello accesso',
            'store': 'Store', 'codice_matricola': 'Matricola', 'ruolo': 'Ruolo',
            'foto': 'Foto profilo', 'commento_mapping': 'Commento mapping',
            'long_absence': 'Long absence', 'is_active': 'Stato account',
        }
        RUOLO_LEVEL = {
            'Retail Assistant': 1,
            'Team Manager': 2, 'Team Manager Visual': 2,
            'Department Manager': 3, 'Visual Manager': 3,
            'Assistant Manager': 4,
            'Store Manager': 5,
        }

        instance = serializer.instance
        old = {}
        for field in serializer.validated_data:
            if field == 'password':
                continue
            if field in ('store', 'ruolo'):
                old[field] = getattr(instance, f'{field}_id', None)
            else:
                old[field] = getattr(instance, field, None)
        old_ruolo_nome = instance.ruolo.nome if instance.ruolo else ''
        old_store_nome = instance.store.nome if instance.store else ''

        instance = serializer.save()

        changed = []
        for field, new_val in serializer.validated_data.items():
            if field == 'password':
                changed.append('Password')
                continue
            if field in ('store', 'ruolo'):
                new_id = new_val.id if new_val else None
                if old.get(field) != new_id:
                    changed.append(FIELD_LABELS.get(field, field))
            else:
                if old.get(field) != new_val:
                    changed.append(FIELD_LABELS.get(field, field))

        detail = f"Modificato: {', '.join(changed)}" if changed else 'Nessuna modifica'
        log_audit(self.request.user, 'modifica', 'utente', instance, detail)

        today = timezone.now().date()

        # Auto-create percorso when ruolo changes
        new_ruolo = serializer.validated_data.get('ruolo')
        if 'ruolo' in serializer.validated_data and (new_ruolo.id if new_ruolo else None) != old.get('ruolo'):
            new_ruolo_nome = new_ruolo.nome if new_ruolo else ''
            old_level = RUOLO_LEVEL.get(old_ruolo_nome, 0)
            new_level = RUOLO_LEVEL.get(new_ruolo_nome, 0)
            if new_level > old_level:
                tipo = 'promozione'
            else:
                tipo = 'cambio_ruolo'
            desc = f'Da {old_ruolo_nome} a {new_ruolo_nome}' if old_ruolo_nome else new_ruolo_nome
            PercorsoCrescita.objects.create(
                user=instance,
                data=today,
                tipo_evento=tipo,
                ruolo_nome=new_ruolo_nome,
                store_nome=instance.store.nome if instance.store else '',
                descrizione=desc,
                created_by=self.request.user,
            )

        # Auto-create percorso when store changes
        new_store = serializer.validated_data.get('store')
        if 'store' in serializer.validated_data and (new_store.id if new_store else None) != old.get('store'):
            new_store_nome = new_store.nome if new_store else ''
            desc = f'Trasferimento da {old_store_nome} a {new_store_nome}' if old_store_nome else new_store_nome
            PercorsoCrescita.objects.create(
                user=instance,
                data=today,
                tipo_evento='trasferimento',
                ruolo_nome=instance.ruolo.nome if instance.ruolo else '',
                store_nome=new_store_nome,
                descrizione=desc,
                created_by=self.request.user,
            )
            # Notify destination store when a store user moves someone to another store
            if self.request.user.livello_accesso == 'store' and new_store and new_store != self.request.user.store:
                NotificaTrasferimento.objects.create(
                    utente_trasferito=instance,
                    store_origine=self.request.user.store,
                    store_destinazione=new_store,
                    snap_nome_utente=f'{instance.cognome} {instance.nome}',
                    snap_nome_store_origine=old_store_nome or (self.request.user.store.nome if self.request.user.store else ''),
                )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False
        instance.save()
        log_audit(request.user, 'disattiva', 'utente', instance,
                  f'Utente disattivato: {instance.email}')
        return Response(status=status.HTTP_204_NO_CONTENT)


class DisattivatIListView(generics.ListAPIView):
    permission_classes = [IsAdmin]
    serializer_class = UserListSerializer

    def get_queryset(self):
        return User.objects.filter(is_active=False).select_related('store', 'ruolo')


class RoleListCreateView(generics.ListCreateAPIView):
    serializer_class = RoleSerializer
    queryset = Role.objects.all()

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated()]
        return [IsAdminOrHO()]


class RoleDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminOrHO]
    serializer_class = RoleSerializer
    queryset = Role.objects.all()


class PercorsoCrescitaListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAdminOrHOOrArea]
    serializer_class = PercorsoCrescitaSerializer

    def get_queryset(self):
        user_id = self.kwargs.get('user_id')
        return PercorsoCrescita.objects.filter(user_id=user_id).select_related('created_by')


class PercorsoCrescitaDeleteView(APIView):
    permission_classes = [IsAdminOrHOOrArea]

    def delete(self, request, pk):
        from datetime import timedelta
        from stores.models import Store
        try:
            entry = PercorsoCrescita.objects.select_related('user', 'user__ruolo', 'user__store').get(pk=pk)
        except PercorsoCrescita.DoesNotExist:
            return Response(status=404)
        if timezone.now() - entry.created_at > timedelta(days=15):
            return Response({'detail': 'Eliminazione consentita solo entro 15 giorni dalla creazione.'}, status=403)
        latest = PercorsoCrescita.objects.filter(user=entry.user).order_by('-data', '-created_at').first()
        if not latest or latest.pk != entry.pk:
            return Response({'detail': 'Puoi eliminare solo l\'ultima variazione.'}, status=403)

        target_user = entry.user
        revert_fields = []

        if entry.tipo_evento in ('promozione', 'cambio_ruolo'):
            # Previous entry of the same type gives us the old role name
            prev = PercorsoCrescita.objects.filter(
                user=target_user, tipo_evento__in=('promozione', 'cambio_ruolo')
            ).exclude(pk=pk).order_by('-data', '-created_at').first()

            old_role_name = prev.ruolo_nome if prev else None

            # Fallback: parse description "Da X a Y" (role names are safe to split)
            if not old_role_name and entry.descrizione.startswith('Da ') and ' a ' in entry.descrizione:
                old_role_name = entry.descrizione[3:].split(' a ', 1)[0]

            if old_role_name:
                try:
                    old_role = Role.objects.get(nome=old_role_name)
                    target_user.ruolo = old_role
                    revert_fields.append('ruolo')
                except Role.DoesNotExist:
                    pass

        elif entry.tipo_evento == 'trasferimento':
            # Previous transfer entry gives us the old store name
            prev = PercorsoCrescita.objects.filter(
                user=target_user, tipo_evento='trasferimento'
            ).exclude(pk=pk).order_by('-data', '-created_at').first()

            old_store_name = prev.store_nome if prev else None

            if old_store_name:
                try:
                    old_store = Store.objects.get(nome=old_store_name)
                    target_user.store = old_store
                    revert_fields.append('store')
                except Store.DoesNotExist:
                    pass

        if revert_fields:
            target_user.save(update_fields=revert_fields)

        detail = f'Eliminato percorso: {TIPO_LABEL_AUDIT.get(entry.tipo_evento, entry.tipo_evento)}'
        if entry.descrizione:
            detail += f' — {entry.descrizione}'
        entry.delete()
        log_audit(request.user, 'elimina', 'utente', target_user, detail)
        return Response(status=204)


class DevelopmentSearchView(APIView):
    permission_classes = [IsAdminOrHOOrArea]

    def get(self, request):
        qs = User.objects.filter(is_active=True).select_related('store', 'ruolo')
        nome = request.query_params.get('nome', '').strip()
        ruolo_id = request.query_params.get('ruolo', '').strip()
        store_id = request.query_params.get('store', '').strip()

        if not (nome or ruolo_id or store_id):
            return Response([])

        if nome:
            qs = qs.filter(Q(cognome__icontains=nome) | Q(nome__icontains=nome))
        if ruolo_id:
            qs = qs.filter(ruolo_id=ruolo_id)
        if store_id:
            qs = qs.filter(store_id=store_id)

        serializer = UserListSerializer(qs[:50], many=True)
        return Response(serializer.data)


class UserDevelopmentDetailView(APIView):
    permission_classes = [IsAdminOrHOOrArea]

    def get(self, request, pk):
        try:
            user = User.objects.select_related('store', 'ruolo').get(pk=pk)
        except User.DoesNotExist:
            return Response({'detail': 'Non trovato.'}, status=404)
        percorso = PercorsoCrescita.objects.filter(user=user).select_related('created_by').order_by('-data')

        # Activity participations
        iscrizioni = (
            Iscrizione.objects
            .filter(user=user)
            .select_related('evento', 'evento__attivita', 'evento__host')
            .order_by('-evento__data', '-evento__ora_inizio')
        )
        attivita_list = [
            {
                'id': i.id,
                'evento_id': i.evento_id,
                'data': i.evento.data.isoformat(),
                'ora_inizio': str(i.evento.ora_inizio)[:5],
                'ore_totali': float(i.evento.ore_totali) if i.evento.ore_totali else None,
                'attivita_nome': i.evento.attivita.nome,
                'attivita_tipologia': i.evento.attivita.tipologia,
                'modalita': i.evento.modalita_partecipazione,
                'host': i.evento.host.descrizione if i.evento.host else '',
                'stato': i.stato,
            }
            for i in iscrizioni
        ]

        return Response({
            'user': UserDetailSerializer(user).data,
            'percorso': PercorsoCrescitaSerializer(percorso, many=True).data,
            'attivita': attivita_list,
        })


class AuditLogListView(generics.ListAPIView):
    permission_classes = [IsAdminOrHO]
    serializer_class = AuditLogSerializer
    pagination_class = None

    def get_queryset(self):
        p = self.request.query_params
        qs = AuditLog.objects.select_related('actor').order_by('-created_at')
        if p.get('tipo'):
            qs = qs.filter(target_tipo=p['tipo'])
        if p.get('azione'):
            qs = qs.filter(azione=p['azione'])
        if p.get('search'):
            qs = qs.filter(
                Q(actor__cognome__icontains=p['search']) |
                Q(actor__nome__icontains=p['search']) |
                Q(target_repr__icontains=p['search']) |
                Q(dettaglio__icontains=p['search'])
            )
        if p.get('data_da'):
            qs = qs.filter(created_at__date__gte=p['data_da'])
        if p.get('data_a'):
            qs = qs.filter(created_at__date__lte=p['data_a'])
        return qs[:500]


def _compute_stats(query_params):
    from events.models import Evento
    from participants.models import Iscrizione
    from django.db.models import Sum, Count, FloatField
    from django.db.models.functions import Coalesce
    from collections import defaultdict

    attivita_id = query_params.get('attivita')
    tipologia = query_params.get('tipologia')
    data_da = query_params.get('data_da')
    data_a = query_params.get('data_a')

    eventi_qs = Evento.objects.select_related('attivita').all()
    if attivita_id:
        eventi_qs = eventi_qs.filter(attivita_id=attivita_id)
    if tipologia:
        eventi_qs = eventi_qs.filter(attivita__tipologia=tipologia)
    if data_da:
        eventi_qs = eventi_qs.filter(data__gte=data_da)
    if data_a:
        eventi_qs = eventi_qs.filter(data__lte=data_a)

    ore_q = list(
        eventi_qs
        .values('attivita_id', 'attivita__nome')
        .annotate(
            ore=Coalesce(Sum('ore_totali'), 0, output_field=FloatField()),
            edizioni=Count('id'),
        )
    )

    def part_by_sesso(sesso):
        return dict(
            Iscrizione.objects.filter(
                evento__in=eventi_qs, stato='partecipato', user__sesso=sesso
            ).values('evento__attivita_id')
            .annotate(n=Count('id'))
            .values_list('evento__attivita_id', 'n')
        )

    def mod_count(mod):
        return dict(
            eventi_qs.filter(modalita_partecipazione=mod)
            .values('attivita_id').annotate(n=Count('id'))
            .values_list('attivita_id', 'n')
        )

    maschi_map   = part_by_sesso('M')
    femmine_map  = part_by_sesso('F')
    ns_map       = part_by_sesso('NS')
    presenza_map = mod_count('presenza')
    online_map   = mod_count('online')
    blended_map  = mod_count('blended')

    # Person-hours per gender: ore_evento * n_partecipanti_M/F
    event_data = {e.id: (float(e.ore_totali or 0), e.attivita_id) for e in eventi_qs}
    ore_maschi_map = defaultdict(float)
    ore_femmine_map = defaultdict(float)
    for g in (Iscrizione.objects.filter(
        evento__in=eventi_qs, stato='partecipato', user__sesso__in=['M', 'F']
    ).values('evento_id', 'user__sesso').annotate(n=Count('id'))):
        ore_ev, att_id = event_data.get(g['evento_id'], (0, None))
        if att_id is None:
            continue
        if g['user__sesso'] == 'M':
            ore_maschi_map[att_id] += ore_ev * g['n']
        else:
            ore_femmine_map[att_id] += ore_ev * g['n']

    result = []
    for row in ore_q:
        att_id = row['attivita_id']
        result.append({
            'attivita_id': att_id,
            'attivita_nome': row['attivita__nome'],
            'num_edizioni': row['edizioni'],
            'ore_erogate': round(float(row['ore']), 2),
            'maschi': maschi_map.get(att_id, 0),
            'ore_maschi': round(ore_maschi_map.get(att_id, 0), 2),
            'femmine': femmine_map.get(att_id, 0),
            'ore_femmine': round(ore_femmine_map.get(att_id, 0), 2),
            'ns': ns_map.get(att_id, 0),
            'presenza': presenza_map.get(att_id, 0),
            'online': online_map.get(att_id, 0),
            'blended': blended_map.get(att_id, 0),
        })

    result.sort(key=lambda x: x['attivita_nome'])

    totale = {
        'num_edizioni': sum(r['num_edizioni'] for r in result),
        'ore_erogate': round(sum(r['ore_erogate'] for r in result), 2),
        'maschi': sum(r['maschi'] for r in result),
        'ore_maschi': round(sum(r['ore_maschi'] for r in result), 2),
        'femmine': sum(r['femmine'] for r in result),
        'ore_femmine': round(sum(r['ore_femmine'] for r in result), 2),
        'ns': sum(r['ns'] for r in result),
        'presenza': sum(r['presenza'] for r in result),
        'online': sum(r['online'] for r in result),
        'blended': sum(r['blended'] for r in result),
    }
    return result, totale


class StatsView(APIView):
    """GET /stats/ — consuntivi attività per admin/HO/area."""
    permission_classes = [IsAdminOrHOOrArea]

    def get(self, request):
        result, totale = _compute_stats(request.query_params)
        return Response({'totale': totale, 'per_attivita': result})


class StatsExportView(APIView):
    """GET /stats/export/ — scarica Excel con i consuntivi."""
    permission_classes = [IsAdminOrHOOrArea]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        result, totale = _compute_stats(request.query_params)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stats Formazione'

        headers = [
            'Attività', 'Edizioni', 'Ore Erogate',
            'Maschi', 'Ore M', 'Femmine', 'Ore F',
            'N.S.', 'In Presenza', 'Online', 'Blended',
        ]
        ws.append(headers)
        header_fill = PatternFill('solid', fgColor='1D4ED8')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for r in result:
            ws.append([
                r['attivita_nome'], r['num_edizioni'], r['ore_erogate'],
                r['maschi'], r['ore_maschi'], r['femmine'], r['ore_femmine'],
                r['ns'] or 0, r['presenza'], r['online'], r['blended'],
            ])

        # Totals row
        tot_row = [
            'TOTALE', totale['num_edizioni'], totale['ore_erogate'],
            totale['maschi'], totale['ore_maschi'], totale['femmine'], totale['ore_femmine'],
            totale['ns'], totale['presenza'], totale['online'], totale['blended'],
        ]
        ws.append(tot_row)
        last_row = ws.max_row
        tot_fill = PatternFill('solid', fgColor='E5E7EB')
        tot_font = Font(bold=True)
        for cell in ws[last_row]:
            cell.fill = tot_fill
            cell.font = tot_font

        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 14

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="stats_formazione.xlsx"'
        wb.save(response)
        return response


class NextMatricolaView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Max
        matricole = (
            User.objects
            .exclude(codice_matricola='')
            .values_list('codice_matricola', flat=True)
        )
        max_num = 0
        for m in matricole:
            try:
                max_num = max(max_num, int(m))
            except (ValueError, TypeError):
                pass
        next_num = max_num + 1
        codice = str(next_num).zfill(5)
        return Response({
            'codice_matricola': codice,
            'email': f'{codice}@kollegas.it',
        })


class RichiestaCreazionoProfiloListCreateView(generics.ListCreateAPIView):
    serializer_class = RichiestaCreazionoProfiloSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        if user.livello_accesso in ('admin', 'ho'):
            return RichiestaCreazioneProfilo.objects.select_related(
                'richiedente', 'richiedente__store', 'store', 'ruolo', 'processed_by'
            ).filter(visto_ho=False)
        if user.livello_accesso == 'store':
            return RichiestaCreazioneProfilo.objects.select_related(
                'richiedente', 'richiedente__store', 'store', 'ruolo', 'processed_by'
            ).filter(richiedente=user, visto_store=False)
        return RichiestaCreazioneProfilo.objects.none()

    def perform_create(self, serializer):
        from rest_framework.exceptions import PermissionDenied, ValidationError
        if self.request.user.livello_accesso != 'store':
            raise PermissionDenied('Solo gli utenti Store possono inviare richieste di creazione profilo.')
        if serializer.validated_data.get('livello_accesso', 'base') != 'base':
            raise ValidationError({'livello_accesso': 'Gli store possono richiedere solo profili Base.'})
        if not serializer.validated_data.get('store'):
            raise ValidationError({'store': 'Il workplace è obbligatorio.'})
        serializer.save(richiedente=self.request.user)


class RichiestaCreazionoProfiloDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return RichiestaCreazioneProfilo.objects.select_related(
                'richiedente', 'store', 'ruolo'
            ).get(pk=pk)
        except RichiestaCreazioneProfilo.DoesNotExist:
            return None

    def patch(self, request, pk):
        richiesta = self.get_object(pk)
        if not richiesta:
            return Response({'detail': 'Non trovata.'}, status=404)

        livello = request.user.livello_accesso

        # Azione "visto": store segna la propria, admin/HO segna la propria
        if request.data.get('visto'):
            if livello in ('admin', 'ho'):
                richiesta.visto_ho = True
                richiesta.save(update_fields=['visto_ho'])
            elif livello == 'store' and richiesta.richiedente == request.user:
                richiesta.visto_store = True
                richiesta.save(update_fields=['visto_store'])
            else:
                return Response({'detail': 'Non autorizzato.'}, status=403)
            return Response({'ok': True})

        # Azione approvazione/rifiuto: solo admin/HO
        if livello not in ('admin', 'ho'):
            return Response({'detail': 'Non autorizzato.'}, status=403)

        if richiesta.stato != 'pending':
            return Response({'detail': 'Richiesta già processata.'}, status=400)

        nuovo_stato = request.data.get('stato')
        if nuovo_stato not in ('approvata', 'rifiutata'):
            return Response({'detail': 'Stato non valido. Usare "approvata" o "rifiutata".'}, status=400)

        if nuovo_stato == 'approvata':
            if User.objects.filter(email=richiesta.email).exists():
                return Response({'detail': f'Email {richiesta.email} già registrata.'}, status=400)
            new_user = User(
                cognome=richiesta.cognome,
                nome=richiesta.nome,
                email=richiesta.email,
                sesso=richiesta.sesso,
                livello_accesso=richiesta.livello_accesso,
                store=richiesta.store,
                codice_matricola=richiesta.codice_matricola,
                ruolo=richiesta.ruolo,
                commento_mapping=richiesta.commento_mapping,
            )
            default_pwd = 'Primark01!'
            new_user.set_password(default_pwd)
            new_user.raw_password = default_pwd
            new_user.save()
            log_audit(request.user, 'crea', 'utente', new_user,
                      f'Utente creato da richiesta store #{richiesta.id}: {new_user.email}')

        richiesta.stato = nuovo_stato
        richiesta.processed_by = request.user
        richiesta.processed_at = timezone.now()
        richiesta.save()
        return Response(RichiestaCreazionoProfiloSerializer(richiesta).data)


class RichiestaEliminazioneProfiloListCreateView(generics.ListCreateAPIView):
    serializer_class = RichiestaEliminazioneProfiloSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        if user.livello_accesso in ('admin', 'ho'):
            return RichiestaEliminazioneProfilo.objects.select_related(
                'target_user', 'richiedente', 'richiedente__store', 'processed_by'
            ).filter(visto_ho=False)
        if user.livello_accesso == 'store':
            return RichiestaEliminazioneProfilo.objects.select_related(
                'target_user', 'richiedente', 'richiedente__store', 'processed_by'
            ).filter(richiedente=user, visto_store=False)
        return RichiestaEliminazioneProfilo.objects.none()

    def perform_create(self, serializer):
        from rest_framework.exceptions import PermissionDenied
        if self.request.user.livello_accesso != 'store':
            raise PermissionDenied('Solo gli utenti Store possono inviare richieste di eliminazione profilo.')
        target = serializer.validated_data['target_user']
        serializer.save(
            richiedente=self.request.user,
            snap_nome=f'{target.cognome} {target.nome}',
        )


class RichiestaEliminazioneProfiloDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return RichiestaEliminazioneProfilo.objects.select_related('target_user').get(pk=pk)
        except RichiestaEliminazioneProfilo.DoesNotExist:
            return None

    def patch(self, request, pk):
        richiesta = self.get_object(pk)
        if not richiesta:
            return Response({'detail': 'Non trovata.'}, status=404)
        livello = request.user.livello_accesso

        if request.data.get('visto'):
            if livello in ('admin', 'ho'):
                richiesta.visto_ho = True
                richiesta.save(update_fields=['visto_ho'])
            elif livello == 'store' and richiesta.richiedente == request.user:
                richiesta.visto_store = True
                richiesta.save(update_fields=['visto_store'])
            else:
                return Response({'detail': 'Non autorizzato.'}, status=403)
            return Response({'ok': True})

        if livello not in ('admin', 'ho'):
            return Response({'detail': 'Non autorizzato.'}, status=403)
        if richiesta.stato != 'pending':
            return Response({'detail': 'Richiesta già processata.'}, status=400)

        nuovo_stato = request.data.get('stato')
        if nuovo_stato not in ('approvata', 'rifiutata'):
            return Response({'detail': 'Stato non valido.'}, status=400)

        if nuovo_stato == 'approvata':
            target = richiesta.target_user
            target.is_active = False
            target.save(update_fields=['is_active'])
            log_audit(request.user, 'disattiva', 'utente', target,
                      f'Disattivato su richiesta store #{richiesta.id}')

        richiesta.stato = nuovo_stato
        richiesta.processed_by = request.user
        richiesta.processed_at = timezone.now()
        richiesta.save()
        return Response(RichiestaEliminazioneProfiloSerializer(richiesta).data)


class AccountBadgeView(APIView):
    """GET /account-badge/ — conta richieste pending per il puntino rosso in sidebar."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        livello = request.user.livello_accesso
        if livello not in ('admin', 'ho'):
            return Response({'count': 0})
        count = (
            RichiestaCreazioneProfilo.objects.filter(stato='pending').count() +
            RichiestaEliminazioneProfilo.objects.filter(stato='pending').count()
        )
        return Response({'count': count})


class NotificaTrasferimentoView(APIView):
    """GET /notifiche-trasferimento/ — notifiche non lette per lo store corrente."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.livello_accesso != 'store' or not request.user.store:
            return Response([])
        qs = NotificaTrasferimento.objects.filter(
            store_destinazione=request.user.store,
            letta=False,
        ).order_by('created_at')
        data = [
            {
                'id': n.id,
                'snap_nome_utente': n.snap_nome_utente,
                'snap_nome_store_origine': n.snap_nome_store_origine,
            }
            for n in qs
        ]
        return Response(data)

    def patch(self, request, pk):
        if request.user.livello_accesso != 'store' or not request.user.store:
            return Response({'detail': 'Non autorizzato.'}, status=403)
        try:
            n = NotificaTrasferimento.objects.get(pk=pk, store_destinazione=request.user.store)
        except NotificaTrasferimento.DoesNotExist:
            return Response({'detail': 'Non trovata.'}, status=404)
        n.letta = True
        n.save(update_fields=['letta'])
        return Response({'ok': True})


class UserDocListCreateView(APIView):
    """GET/POST /users/<user_id>/docs/"""
    permission_classes = [IsAuthenticated]

    def _can_access(self, request, target_user_id):
        u = request.user
        if u.livello_accesso in ('admin', 'ho'):
            return True
        if u.livello_accesso == 'store':
            try:
                target = User.objects.get(pk=target_user_id)
                return target.store_id == u.store_id
            except User.DoesNotExist:
                pass
        return False

    def get(self, request, user_id):
        if not self._can_access(request, user_id):
            return Response({'detail': 'Non autorizzato.'}, status=403)
        docs = UserDocument.objects.filter(user_id=user_id)
        return Response([
            {'id': d.id, 'nome_file': d.nome_file, 'created_at': d.created_at}
            for d in docs
        ])

    def post(self, request, user_id):
        u = request.user
        if u.livello_accesso not in ('store', 'admin', 'ho'):
            return Response({'detail': 'Non autorizzato.'}, status=403)
        if not self._can_access(request, user_id):
            return Response({'detail': 'Non autorizzato.'}, status=403)
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'detail': 'File mancante.'}, status=400)
        nome = request.data.get('nome_file') or uploaded.name
        doc = UserDocument.objects.create(
            user_id=user_id,
            nome_file=nome,
            file=uploaded,
            uploaded_by=u,
        )
        return Response({'id': doc.id, 'nome_file': doc.nome_file, 'created_at': doc.created_at}, status=201)


class UserDocDetailView(APIView):
    """DELETE /docs/<pk>/  |  GET /docs/<pk>/ (serve file)"""
    permission_classes = [IsAuthenticated]

    def _get_doc(self, pk):
        try:
            return UserDocument.objects.select_related('user').get(pk=pk)
        except UserDocument.DoesNotExist:
            return None

    def _can_access(self, request, doc):
        u = request.user
        if u.livello_accesso in ('admin', 'ho'):
            return True
        if u.livello_accesso == 'store':
            return doc.user.store_id == u.store_id
        return False

    def get(self, request, pk):
        from django.http import FileResponse
        doc = self._get_doc(pk)
        if not doc:
            return Response({'detail': 'Non trovato.'}, status=404)
        if not self._can_access(request, doc):
            return Response({'detail': 'Non autorizzato.'}, status=403)
        response = FileResponse(doc.file.open('rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{doc.nome_file}"'
        return response

    def delete(self, request, pk):
        u = request.user
        if u.livello_accesso in ('admin', 'ho'):
            return Response({'detail': 'Non autorizzato a eliminare.'}, status=403)
        doc = self._get_doc(pk)
        if not doc:
            return Response({'detail': 'Non trovato.'}, status=404)
        if not self._can_access(request, doc):
            return Response({'detail': 'Non autorizzato.'}, status=403)
        doc.file.delete(save=False)
        doc.delete()
        return Response(status=204)
