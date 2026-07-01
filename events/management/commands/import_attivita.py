import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from events.models import AttivitaCatalogo

TIPO_REPORT_MAP = {
    'corso di formazione':     'formazione',
    'attività interna store':  'sviluppo',
    'attivita interna store':  'sviluppo',
}


def infer_tipologia(nome):
    n = nome.lower()
    if any(k in n for k in [
        'gwu ', 'grow with us', 'assessment', 'development',
        'growth program', 'leadership', 'mentoring', 'tm2be',
        'tm-dm', 'dm-am', 'our leadership', 'talent university',
        'connection week',
    ]):
        return 'sviluppo'
    if any(k in n for k in ['d&i', 'ehs', 'compliance', 'policy e procedure']):
        return 'compliance'
    if any(k in n for k in ['weareprimark', 'primark x tutti']):
        return 'onboarding'
    if n in ['visita', 'popsquare', 'area manager growth program']:
        return 'altro'
    return 'formazione'


class Command(BaseCommand):
    help = 'Importa attività da catalogo attività.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?',
            default=r'C:\Users\matti\OneDrive\Desktop\Popsquare2.0\catalogo attività.xlsx')
        parser.add_argument('--report',
            default=r'C:\Users\matti\OneDrive\Desktop\Popsquare2.0\report.xlsx',
            help='Usa il report per ricavare la tipologia reale')

    def handle(self, *args, **options):
        # Ricava tipologie reali dal report
        tipo_da_report = {}
        wb_rep = openpyxl.load_workbook(options['report'])
        for row in wb_rep.active.iter_rows(min_row=2, values_only=True):
            _, att, _, tipo, *_ = row
            if att and tipo:
                nome_att = str(att).strip()
                tipo_db = TIPO_REPORT_MAP.get(str(tipo).strip().lower())
                if tipo_db and nome_att not in tipo_da_report:
                    tipo_da_report[nome_att] = tipo_db

        wb = openpyxl.load_workbook(options['file'])
        ws = wb.active

        creati = gia_presenti = 0
        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                nome = str(row[0]).strip()
                tipologia = tipo_da_report.get(nome) or infer_tipologia(nome)

                _, created = AttivitaCatalogo.objects.get_or_create(
                    nome=nome,
                    defaults={'tipologia': tipologia, 'is_active': True}
                )
                if created:
                    creati += 1
                    self.stdout.write(f'  + [{tipologia:12s}] {nome}')
                else:
                    gia_presenti += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nAttività: {creati} create, {gia_presenti} già presenti'
        ))
